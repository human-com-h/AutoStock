from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.ulid import new_ulid
from app.models.master_data import Part
from app.models.orders import PurchaseItem, PurchaseOrder, SalesItem, SalesOrder
from app.models.stock import StockLedger
from app.services.part_service import create_part, search_parts
from app.services.stock_service import append_ledger_entry, list_inventory

_ORDER_TYPE_LABELS = {
    "purchase": "采购入库",
    "purchase_return": "采购退货",
    "sale": "销售出库",
    "sale_return": "销售退货",
}


def _order_type_label(value: str) -> str:
    return _ORDER_TYPE_LABELS.get(value, "其他业务")


PART_HEADERS = [
    "零件编号",
    "OE号",
    "零件名称",
    "规格",
    "单位",
    "参考进价(分)",
    "参考售价(分)",
    "最低库存",
    "最高库存",
    "货位",
    "适用车型",
    "期初库存",
]


def _workbook_bytes(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def part_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "零件导入"
    sheet.append(PART_HEADERS)
    sheet.append(
        [
            "P001",
            "OE-001,OE-002",
            "机油滤清器",
            "标准",
            "个",
            1200,
            1800,
            2,
            50,
            "A-01",
            "通用",
            10,
        ]
    )
    sheet.freeze_panes = "A2"
    return _workbook_bytes(workbook)


def export_parts(db: Session) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "零件档案"
    sheet.append(PART_HEADERS[:-1])
    for row in search_parts(db, limit=10000):
        sheet.append(
            [
                row.part_number,
                row.oe_number,
                row.name,
                row.spec,
                row.unit,
                row.purchase_price,
                row.sale_price,
                float(row.min_stock),
                float(row.max_stock) if row.max_stock is not None else None,
                row.location,
                row.vehicle_models,
            ]
        )
    sheet.freeze_panes = "A2"
    return _workbook_bytes(workbook)


def export_inventory(db: Session) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "库存"
    sheet.append(["零件编号", "零件名称", "货位", "库存", "单位", "平均成本(分)", "库存金额(分)"])
    for row in list_inventory(db, limit=10000):
        sheet.append(
            [
                row["part_number"],
                row["name"],
                row["location"],
                row["quantity"],
                row["unit"],
                row["avg_cost"],
                row["stock_amount"],
            ]
        )
    sheet.freeze_panes = "A2"
    return _workbook_bytes(workbook)


def export_orders(db: Session, kind: str) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    if kind == "purchase":
        sheet.title = "采购明细"
        sheet.append(
            ["单号", "日期", "类型", "零件编号", "零件名称", "数量", "进价(分)", "金额(分)"]
        )
        rows = db.execute(
            select(PurchaseOrder, PurchaseItem, Part)
            .join(PurchaseItem, PurchaseItem.order_id == PurchaseOrder.id)
            .join(Part, Part.id == PurchaseItem.part_id)
            .where(PurchaseOrder.is_deleted == 0, PurchaseItem.is_deleted == 0)
            .order_by(PurchaseOrder.order_date.desc(), PurchaseOrder.order_no.desc())
        )
        for order, item, part in rows:
            sheet.append(
                [
                    order.order_no,
                    order.order_date,
                    _order_type_label(order.order_type),
                    part.part_number,
                    part.name,
                    float(item.quantity),
                    item.purchase_price,
                    item.amount,
                ]
            )
    elif kind == "sales":
        sheet.title = "销售明细"
        sheet.append(
            [
                "单号",
                "日期",
                "类型",
                "客户",
                "零件编号",
                "零件名称",
                "数量",
                "售价(分)",
                "金额(分)",
                "成本(分)",
                "毛利(分)",
            ]
        )
        rows = db.execute(
            select(SalesOrder, SalesItem, Part)
            .join(SalesItem, SalesItem.order_id == SalesOrder.id)
            .join(Part, Part.id == SalesItem.part_id)
            .where(SalesOrder.is_deleted == 0, SalesItem.is_deleted == 0)
            .order_by(SalesOrder.order_date.desc(), SalesOrder.order_no.desc())
        )
        for order, item, part in rows:
            sheet.append(
                [
                    order.order_no,
                    order.order_date,
                    _order_type_label(order.order_type),
                    order.customer_name,
                    part.part_number,
                    part.name,
                    float(item.quantity),
                    item.sale_price,
                    item.amount,
                    item.cost_amount,
                    item.amount - item.cost_amount,
                ]
            )
    else:
        raise ValueError("未知单据导出类型")
    sheet.freeze_panes = "A2"
    return _workbook_bytes(workbook)


def export_stock_ledger(db: Session) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "库存台账"
    sheet.append(
        [
            "发生时间",
            "零件编号",
            "零件名称",
            "变动类型",
            "数量",
            "单位成本(分)",
            "来源类型",
            "来源ID",
        ]
    )
    rows = db.execute(
        select(StockLedger, Part)
        .join(Part, Part.id == StockLedger.part_id)
        .order_by(StockLedger.occurred_at, StockLedger.rev)
    )
    for ledger, part in rows:
        sheet.append(
            [
                ledger.occurred_at,
                part.part_number,
                part.name,
                ledger.change_type,
                float(ledger.quantity),
                ledger.unit_cost,
                ledger.source_type,
                ledger.source_id,
            ]
        )
    sheet.freeze_panes = "A2"
    return _workbook_bytes(workbook)


def export_stock_summary(db: Session) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "进销存汇总"
    sheet.append(
        [
            "零件编号",
            "零件名称",
            "期初库存",
            "采购入库",
            "销售退回",
            "销售出库",
            "采购退货",
            "盘点调整",
            "期末库存",
            "平均成本(分)",
            "库存金额(分)",
        ]
    )
    changes: dict[str, dict[str, Decimal]] = {}
    for part_id, change_type, quantity in db.execute(
        select(StockLedger.part_id, StockLedger.change_type, StockLedger.quantity)
    ):
        row = changes.setdefault(part_id, {})
        row[change_type] = row.get(change_type, Decimal("0")) + Decimal(str(quantity))

    for row in list_inventory(db, limit=10000):
        values = changes.get(row["part_id"], {})
        sheet.append(
            [
                row["part_number"],
                row["name"],
                float(values.get("opening", 0)),
                float(values.get("purchase", 0)),
                float(values.get("sale_return", 0)),
                float(-values.get("sale", 0)),
                float(-values.get("purchase_return", 0)),
                float(values.get("adjust", 0)),
                row["quantity"],
                row["avg_cost"],
                row["stock_amount"],
            ]
        )
    sheet.freeze_panes = "A2"
    return _workbook_bytes(workbook)


def import_parts(db: Session, content: bytes) -> dict:
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    missing = [name for name in ("零件编号", "零件名称") if name not in headers]
    if missing:
        return {"imported": 0, "errors": [{"row": 1, "message": f"缺少列：{','.join(missing)}"}]}
    indexes = {name: index for index, name in enumerate(headers)}
    imported = 0
    errors: list[dict] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value not in (None, "") for value in values):
            continue

        def value(name: str, default=None, row_values=values):
            index = indexes.get(name)
            return (
                row_values[index]
                if index is not None and index < len(row_values)
                else default
            )

        try:
            part = create_part(
                db,
                part_number=str(value("零件编号") or "").strip(),
                oe_number=str(value("OE号") or "").strip() or None,
                name=str(value("零件名称") or "").strip(),
                spec=str(value("规格") or "").strip() or None,
                brand_id=None,
                category_id=None,
                supplier_id=None,
                unit=str(value("单位") or "个").strip(),
                purchase_price=int(value("参考进价(分)", 0) or 0),
                sale_price=int(value("参考售价(分)", 0) or 0),
                min_stock=Decimal(str(value("最低库存", 0) or 0)),
                max_stock=(
                    Decimal(str(value("最高库存")))
                    if value("最高库存") not in (None, "")
                    else None
                ),
                location=str(value("货位") or "").strip() or None,
                vehicle_models=str(value("适用车型") or "").strip() or None,
                remark=None,
            )
            opening = Decimal(str(value("期初库存", 0) or 0))
            if opening:
                append_ledger_entry(
                    db,
                    part_id=part.id,
                    change_type="opening",
                    quantity=opening,
                    source_type="opening_import",
                    source_id=new_ulid(),
                    unit_cost=part.purchase_price,
                )
                db.commit()
            imported += 1
        except Exception as exc:
            db.rollback()
            errors.append({"row": row_number, "message": str(exc)})
    return {"imported": imported, "errors": errors}
