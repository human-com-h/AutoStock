from io import BytesIO

from openpyxl import load_workbook


def _part(client, number="FEATURE-001"):
    response = client.post(
        "/api/parts",
        json={
            "part_number": number,
            "name": "阶段一测试零件",
            "unit": "个",
            "purchase_price": 500,
            "sale_price": 800,
        },
    )
    assert response.status_code == 200
    return response.json()["data"]


def test_settings_apply_to_new_sales_orders(app_client):
    current = app_client.get("/api/settings")
    assert current.status_code == 200
    updated = app_client.put(
        "/api/settings",
        json={
            "shop_name": "测试汽配店",
            "default_unit": "件",
            "allow_negative_stock": False,
            "stale_days": 90,
        },
    )
    assert updated.status_code == 200
    assert updated.json()["data"]["shop_name"] == "测试汽配店"
    assert updated.json()["data"]["print_warehouse"] == "主仓库"
    assert updated.json()["data"]["print_notice"] == "商品如有质量问题，请及时联系我们处理。"
    assert [field["label"] for field in updated.json()["data"]["print_custom_fields"]] == [
        "运输方式",
        "运费承担",
        "物流单号",
    ]

    part = _part(app_client)
    blocked = app_client.post(
        "/api/orders/sales",
        json={"items": [{"part_id": part["id"], "quantity": 1, "sale_price": 800}]},
    )
    assert blocked.status_code == 400
    assert blocked.json()["error"]["code"] == "BUSINESS_STOCK_INSUFFICIENT"


def test_print_template_settings_support_recommended_and_custom_fields(app_client):
    settings = app_client.get("/api/settings").json()["data"]
    settings.update(
        {
            "print_payment_account": "工商银行 6222 **** 1234",
            "print_wechat": "AUTO-PARTS",
            "print_warranty_period": "三包期内凭单退换",
            "print_reviewer": "复核员",
            "print_custom_fields": [
                {
                    "label": "运输方式",
                    "value": "",
                    "visible": True,
                    "handwritten": True,
                },
                {
                    "label": "内部备注",
                    "value": "不显示",
                    "visible": False,
                    "handwritten": False,
                },
            ],
        }
    )

    updated = app_client.put("/api/settings", json=settings)
    assert updated.status_code == 200
    data = updated.json()["data"]
    assert data["print_payment_account"] == "工商银行 6222 **** 1234"
    assert data["print_wechat"] == "AUTO-PARTS"
    assert data["print_warranty_period"] == "三包期内凭单退换"
    assert data["print_reviewer"] == "复核员"
    assert data["print_custom_fields"] == settings["print_custom_fields"]

    too_many = settings.copy()
    too_many["print_custom_fields"] = [
        {
            "label": f"字段{index}",
            "value": "",
            "visible": True,
            "handwritten": False,
        }
        for index in range(6)
    ]
    assert app_client.put("/api/settings", json=too_many).status_code == 422


def test_dashboard_and_rankings_use_real_orders(app_client):
    part = _part(app_client)
    app_client.post(
        "/api/orders/purchases",
        json={"items": [{"part_id": part["id"], "quantity": 10, "purchase_price": 500}]},
    )
    app_client.post(
        "/api/orders/sales",
        json={
            "customer_name": "测试客户",
            "items": [{"part_id": part["id"], "quantity": 2, "sale_price": 800}],
        },
    )
    dashboard = app_client.get("/api/reports/dashboard").json()["data"]
    assert dashboard["today_sales"] == 1600
    assert dashboard["today_profit"] == 600
    assert dashboard["period_sales"] == 1600
    assert dashboard["period_profit"] == 600
    assert dashboard["period_purchase_amount"] == 5000
    assert dashboard["gross_margin"] == 37.5
    assert dashboard["inventory_amount"] == 4000
    assert len(dashboard["trend"]) == 30
    assert dashboard["trend"][-1]["sales"] == 1600
    assert dashboard["recent_orders"][0]["order_no"]

    rankings = app_client.get("/api/reports/rankings").json()["data"]
    assert rankings["parts"][0]["part_id"] == part["id"]
    assert rankings["parts"][0]["sales"] == 1600
    assert rankings["parts"][0]["profit"] == 600
    assert rankings["customers"][0] == {"name": "测试客户", "sales": 1600}


def test_part_stock_history_includes_running_balance_and_print_document(app_client):
    part = _part(app_client)
    purchase = app_client.post(
        "/api/orders/purchases",
        json={"items": [{"part_id": part["id"], "quantity": 10, "purchase_price": 500}]},
    ).json()["data"]
    sale = app_client.post(
        "/api/orders/sales",
        json={"items": [{"part_id": part["id"], "quantity": 2, "sale_price": 800}]},
    ).json()["data"]

    response = app_client.get(f"/api/stock/{part['id']}/history")
    assert response.status_code == 200
    history = response.json()["data"]
    assert history["part"]["part_number"] == part["part_number"]
    assert history["current_quantity"] == 8
    assert history["total"] == 2
    assert [row["quantity"] for row in history["entries"]] == [-2, 10]
    assert [row["balance_after"] for row in history["entries"]] == [8, 10]
    assert history["entries"][0]["document"] == {
        "kind": "sales",
        "id": sale["id"],
        "order_no": sale["order_no"],
        "order_type": "sale",
        "available": True,
    }
    assert history["entries"][1]["document"]["id"] == purchase["id"]
    assert (
        app_client.get(f"/api/orders/sales/{sale['id']}/pdf").headers["content-type"]
        == "application/pdf"
    )


def test_part_stock_history_marks_void_document_unavailable(app_client):
    part = _part(app_client)
    purchase = app_client.post(
        "/api/orders/purchases",
        json={"items": [{"part_id": part["id"], "quantity": 3, "purchase_price": 500}]},
    ).json()["data"]
    voided = app_client.post(f"/api/orders/purchases/{purchase['id']}/void")
    assert voided.status_code == 200

    history = app_client.get(f"/api/stock/{part['id']}/history").json()["data"]
    assert history["current_quantity"] == 0
    assert history["entries"][0]["source_type"].startswith("purchase_item_void")
    assert history["entries"][0]["document"]["order_no"] == purchase["order_no"]
    assert history["entries"][0]["document"]["available"] is False


def test_reconcile_endpoint_detects_snapshot_difference(app_client):
    part = _part(app_client)
    app_client.post(
        "/api/orders/purchases",
        json={"items": [{"part_id": part["id"], "quantity": 10, "purchase_price": 500}]},
    )
    clean = app_client.get("/api/stock/reconcile").json()["data"]
    assert clean == {
        "ok": True,
        "checked_count": 1,
        "mismatch_count": 0,
        "differences": [],
    }

    from app.db.session import SessionLocal
    from app.models.stock import StockSnapshot

    with SessionLocal() as db:
        snapshot = db.get(StockSnapshot, part["id"])
        snapshot.quantity = 9
        db.commit()

    mismatch = app_client.get("/api/stock/reconcile").json()["data"]
    assert mismatch["ok"] is False
    assert mismatch["mismatch_count"] == 1
    assert mismatch["differences"][0]["difference"] == -1


def test_excel_template_import_and_exports(app_client):
    template = app_client.get("/api/excel/template/parts")
    assert template.status_code == 200
    workbook = load_workbook(BytesIO(template.content))
    assert workbook.active["A1"].value == "零件编号"

    imported = app_client.post(
        "/api/excel/import/parts",
        files={
            "file": (
                "parts.xlsx",
                template.content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert imported.status_code == 200
    assert imported.json()["data"]["imported"] == 1
    inventory = app_client.get("/api/stock").json()["data"]
    assert inventory[0]["quantity"] == 10

    parts_export = app_client.get("/api/excel/export/parts")
    inventory_export = app_client.get("/api/excel/export/inventory")
    assert parts_export.content[:2] == b"PK"
    assert inventory_export.content[:2] == b"PK"


def test_order_ledger_and_summary_exports(app_client):
    part = _part(app_client)
    app_client.post(
        "/api/orders/purchases",
        json={"items": [{"part_id": part["id"], "quantity": 10, "purchase_price": 500}]},
    )
    app_client.post(
        "/api/orders/sales",
        json={"items": [{"part_id": part["id"], "quantity": 2, "sale_price": 800}]},
    )
    exports = {
        "purchase": app_client.get("/api/excel/export/orders/purchases"),
        "sales": app_client.get("/api/excel/export/orders/sales"),
        "ledger": app_client.get("/api/excel/export/ledger"),
        "summary": app_client.get("/api/excel/export/summary"),
    }
    assert all(response.status_code == 200 for response in exports.values())
    assert all(response.content[:2] == b"PK" for response in exports.values())

    purchase_export = load_workbook(BytesIO(exports["purchase"].content)).active
    sales_export = load_workbook(BytesIO(exports["sales"].content)).active
    assert purchase_export["C2"].value == "采购入库"
    assert sales_export["C2"].value == "销售出库"

    summary = load_workbook(BytesIO(exports["summary"].content)).active
    assert summary["A2"].value == part["part_number"]
    assert summary["D2"].value == 10
    assert summary["F2"].value == 2
    assert summary["I2"].value == 8


def test_backup_create_list_and_restore_roundtrip(app_client):
    first = _part(app_client, "BACKUP-001")
    created = app_client.post("/api/backups")
    assert created.status_code == 200
    name = created.json()["data"]["name"]
    assert name in {row["name"] for row in app_client.get("/api/backups").json()["data"]}

    _part(app_client, "BACKUP-002")
    from app.services import backup_service

    backup_service.restore_backup(name)

    from app.db.session import SessionLocal
    from app.models.master_data import Part

    with SessionLocal() as db:
        numbers = {row.part_number for row in db.query(Part).all()}
    assert first["part_number"] in numbers
    assert "BACKUP-002" not in numbers
