from __future__ import annotations

import io
import json
import zipfile
from datetime import datetime, timedelta

from openpyxl import load_workbook


def _create_part(client, number: str, name: str = "机油滤清器"):
    response = client.post(
        "/api/parts",
        json={
            "part_number": number,
            "name": name,
            "unit": "个",
            "purchase_price": 1200,
            "sale_price": 1800,
        },
    )
    assert response.status_code == 200
    return response.json()["data"]


def test_master_data_history_can_restore_previous_version(app_client):
    part = _create_part(app_client, "HISTORY-001")
    updated = app_client.put(
        f"/api/parts/{part['id']}",
        json={"name": "新名称", "sale_price": 2300},
    )
    assert updated.status_code == 200

    history = app_client.get("/api/history", params={"entity_type": "part"}).json()["data"]
    update_event = next(row for row in history if row["action"] == "update")
    assert update_event["before"]["name"] == "机油滤清器"
    assert update_event["after"]["name"] == "新名称"
    assert update_event["can_restore"] is True

    restored = app_client.post(
        f"/api/history/{update_event['id']}/restore",
        json={"confirm": "RESTORE"},
    )
    assert restored.status_code == 200
    current = app_client.get(f"/api/parts/{part['id']}").json()["data"]
    assert current["name"] == "机油滤清器"
    assert current["sale_price"] == 1800
    assert app_client.get("/api/stock/reconcile").json()["data"]["ok"] is True


def test_non_master_history_is_auditable_but_not_directly_restorable(app_client):
    part = _create_part(app_client, "HISTORY-ORDER-001")
    order = app_client.post(
        "/api/orders/purchases",
        json={"items": [{"part_id": part["id"], "quantity": 2, "purchase_price": 1200}]},
    )
    assert order.status_code == 200
    history = app_client.get(
        "/api/history",
        params={"entity_type": "purchase_order"},
    ).json()["data"]
    assert history[0]["action"] == "create"
    assert history[0]["can_restore"] is False


def test_restore_point_keeps_label_summary_and_validation(app_client):
    _create_part(app_client, "POINT-001")
    created = app_client.post("/api/backups", json={"label": "批量导入前"})
    assert created.status_code == 200
    data = created.json()["data"]
    assert data["label"] == "批量导入前"
    assert data["verified"] is True
    assert data["summary"]["parts"] == 1


def test_expired_restore_points_are_removed_by_policy(app_client):
    created = app_client.post("/api/backups", json={"label": "过期测试"}).json()["data"]
    from app.core.config import settings

    metadata_path = (settings.data_dir / "backups" / created["name"]).with_suffix(".json")
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    metadata["created_at"] = (
        datetime.now().astimezone() - timedelta(days=181)
    ).isoformat()
    metadata_path.write_text(json.dumps(metadata, ensure_ascii=False), encoding="utf-8")

    from app.services import backup_service

    removed = backup_service.apply_retention_policy()
    assert created["name"] in removed
    assert created["name"] not in {
        row["name"] for row in app_client.get("/api/backups").json()["data"]
    }


def test_migration_package_is_readable_and_can_roundtrip(app_client):
    original = _create_part(app_client, "MIGRATE-001")
    app_client.post(
        "/api/orders/purchases",
        json={"items": [{"part_id": original["id"], "quantity": 5, "purchase_price": 1200}]},
    )

    exported = app_client.get("/api/backups/migration/export")
    assert exported.status_code == 200
    with zipfile.ZipFile(io.BytesIO(exported.content)) as archive:
        names = set(archive.namelist())
        assert {
            "autostock.db",
            "manifest.json",
            "checksums.txt",
            "README.txt",
            "经营数据.xlsx",
            "csv/零件档案.csv",
            "csv/当前库存.csv",
            "csv/库存流水.csv",
        }.issubset(names)
        workbook = load_workbook(io.BytesIO(archive.read("经营数据.xlsx")), data_only=True)
        assert "零件档案" in workbook.sheetnames
        assert workbook["零件档案"]["A2"].value == "MIGRATE-001"
        assert workbook["当前库存"]["D2"].value == 5

    _create_part(app_client, "MIGRATE-EXTRA")
    imported = app_client.post(
        "/api/backups/migration/import",
        content=exported.content,
        headers={"Content-Type": "application/zip"},
    )
    assert imported.status_code == 200, imported.text
    numbers = {
        row["part_number"]
        for row in app_client.get("/api/parts", params={"limit": 200}).json()["data"]
    }
    assert "MIGRATE-001" in numbers
    assert "MIGRATE-EXTRA" not in numbers
    assert app_client.get("/api/stock/reconcile").json()["data"]["ok"] is True
